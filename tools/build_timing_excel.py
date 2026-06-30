import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

FONT = "Arial"

with open("tests_timing.csv", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

wb = Workbook()
ws = wb.active
ws.title = "Comparativo de Tempo"

title_font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", start_color="1F4E78")
header_font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="2E75B6")
normal_font = Font(name=FONT, size=11)

thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.merge_cells("A1:H1")
ws["A1"] = "Comparação de Tempo de Execução das Suítes de Teste (make timing)"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:H2")
ws["A2"] = (
    "Cada suíte foi executada repetidamente para obter medições estáveis "
    "usando clock_gettime(CLOCK_MONOTONIC)."
)
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="555555")

headers = [
    "Suíte de Teste",
    "Repetições",
    "Testes por execução",
    "Asserções por execução",
    "Falhas",
    "Tempo total (s)",
    "Tempo médio por execução (ms)"
]

header_row = 4
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

data_start = header_row + 1
for i, r in enumerate(rows):
    row = data_start + i
    values = [
        r["Suite"],
        int(r["Repeticoes"]),
        int(r["Testes_por_execucao"]),
        int(r["Assertions_por_execucao"]),
        int(r["Falhas"]),
        float(r["Tempo_total_s"]),
        float(r["Tempo_medio_por_execucao_ms"])
    ]

    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = normal_font
        c.border = border
        c.alignment = Alignment(
            horizontal="left" if col == 1 else "center",
            vertical="center"
        )

    ws.cell(row=row, column=6).number_format = "0.000000"
    ws.cell(row=row, column=7).number_format = "0.000000"

last_row = data_start + len(rows) - 1

rank_row = last_row + 2

ws.cell(row=rank_row, column=1, value="Mais rápida:").font = Font(name=FONT, bold=True)
ws.cell(
    row=rank_row,
    column=2,
    value=f'=INDEX(A{data_start}:A{last_row},MATCH(MIN(G{data_start}:G{last_row}),G{data_start}:G{last_row},0))'
)

rank_row2 = rank_row + 1

ws.cell(row=rank_row2, column=1, value="Mais lenta:").font = Font(name=FONT, bold=True)
ws.cell(
    row=rank_row2,
    column=2,
    value=f'=INDEX(A{data_start}:A{last_row},MATCH(MAX(G{data_start}:G{last_row}),G{data_start}:G{last_row},0))'
)

extra_col = 8
ws.cell(row=header_row, column=extra_col, value="x vs. mais rápida")
ws.cell(row=header_row, column=extra_col).font = header_font
ws.cell(row=header_row, column=extra_col).fill = header_fill
ws.cell(row=header_row, column=extra_col).border = border

for i in range(len(rows)):
    row = data_start + i
    c = ws.cell(
        row=row,
        column=extra_col,
        value=f"=G{row}/MIN(G${data_start}:G${last_row})"
    )
    c.number_format = '0.00"x"'
    c.font = normal_font
    c.border = border
    c.alignment = Alignment(horizontal="center")

widths = [42, 12, 18, 20, 10, 16, 24, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

chart = BarChart()
chart.type = "col"
chart.title = "Tempo médio por execução (ms)"
chart.y_axis.title = "Tempo (ms)"
chart.x_axis.title = "Suíte"
chart.style = 10
chart.width = 24
chart.height = 12

data_ref = Reference(
    ws,
    min_col=7,
    min_row=header_row,
    max_row=last_row
)

cats_ref = Reference(
    ws,
    min_col=1,
    min_row=data_start,
    max_row=last_row
)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.legend = None

ws.add_chart(chart, f"A{rank_row2 + 3}")

wb.save("tests_timing.xlsx")
print("tests_timing.xlsx gerado com sucesso.")