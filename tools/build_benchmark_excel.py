import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

FONT = "Arial"

with open("benchmark_results.csv", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

wb = Workbook()
ws = wb.active
ws.title = "Benchmark"

title_font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", start_color="1F4E78")
header_font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="2E75B6")
normal_font = Font(name=FONT, size=11)

thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.merge_cells("A1:G1")
ws["A1"] = "Benchmark de Hash Tables"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

headers = [
    "Implementação",
    "Carga (alpha)",
    "Inserção (ms)",
    "Busca Hit (ms)",
    "Busca Miss (ms)",
    "Remoção (ms)",
    "Total (ms)"
]

header_row = 3

for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = border

data_start = header_row + 1

for i, r in enumerate(rows):
    row = data_start + i

    insert_t = float(r["Insert_ms"])
    hit_t = float(r["SearchHit_ms"])
    miss_t = float(r["SearchMiss_ms"])
    remove_t = float(r["Delete_ms"])
    total_t = insert_t + hit_t + miss_t + remove_t

    values = [
        r["Implementacao"],
        float(r["Alpha"]),
        insert_t,
        hit_t,
        miss_t,
        remove_t,
        total_t
    ]

    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = normal_font
        c.border = border
        c.alignment = Alignment(horizontal="center")

last_row = data_start + len(rows) - 1

widths = [28, 14, 16, 16, 16, 16, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

chart = BarChart()
chart.title = "Tempo Total por Implementação"
chart.y_axis.title = "Tempo (ms)"
chart.x_axis.title = "Implementação"
chart.style = 10
chart.width = 20
chart.height = 10

data_ref = Reference(
    ws,
    min_col=7,
    min_row=header_row,
    max_row=last_row
)

cats_ref = Reference(
    ws,
    min_col=1,
    min_row=data_start,
    max_row=last_row
)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

ws.add_chart(chart, "I4")

wb.save("benchmark_results.xlsx")
print("benchmark_results.xlsx gerado com sucesso.")